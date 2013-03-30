from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

def workbook_from_aggregation(filename, aggregation):
    wb = Workbook()

    ws = wb.worksheets[0]

    ws.title = "Aggregations"

    for col_idx in xrange(1, 40):
        col = get_column_letter(col_idx)
        for row in xrange(1, 600):
            ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

    wb.save(filename=filename)
    return True
            
